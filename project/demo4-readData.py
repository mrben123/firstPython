from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('../static/excels/new_excel.xlsx')
ws = wb.active

for row in range(1,8):
    for col in range(1,7):
        char = get_column_letter(col)
        print(ws[char + str(row)])