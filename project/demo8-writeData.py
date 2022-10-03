from msilib.schema import Font
from tkinter.font import BOLD
from tokenize import Number
from turtle import color
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# wb_url = '../static/excels/new_excel.xlsx'
data = [
    {
        'name': '小白',
        'tall': 155,
        'age': 21,
        'weight': 120
    },
    {
        'name': '小黑',
        'tall': 160,
        'age': 21,
        'weight': 110
    },
    {
        'name': '小红',
        'tall': 165,
        'age': 21,
        'weight': 140
    },
    {
        'name': '小蓝',
        'tall': 170,
        'age': 21,
        'weight': 130
    }
]

wb = Workbook()
ws = wb.active

title = ['姓名', '身高', '年纪', '体重']
ws.append(title)

# ws.move_range('B2:B2',rows=2)

for person in data:
    temp = list(person.values()) 
    ws.append(temp)

for col in range(1,5):
    char = get_column_letter(col)
    #col like A,B,C,D
    # print(char)
    # print(ws[char + '2'].value)
    # print(type(ws[char + '2'].value))
    print(char)

    ws[char + '1'].font = Font(bold = True, color='00FFFF99')
    print(ws[char + '1'])
    
    #如果row的第二个值是数字才会计算平均值
    if type(ws[char + '2'].value) == int:
        ws[char + '6'] = f"=AVERAGE({char + '2'}:{char + '5'})"

wb.save('../static/excels/test.xlsx')