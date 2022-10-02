from openpyxl import Workbook, load_workbook

wb_url = '../static/excels/new_excel.xlsx'
wb = load_workbook(wb_url)
ws = wb.active

# ws.merge_cells('A8:F6')
ws.unmerge_cells('A8:F8')
wb.save(wb_url)