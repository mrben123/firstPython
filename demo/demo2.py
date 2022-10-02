from openpyxl import workbook, load_workbook
#wordkbook就是excel档案的意思
#load_workbook是读取档案的意思
def main():
    wb = load_workbook('../static/excels/excel_1.xlsx')
    ws = wb.active
    print('当前工作表：', ws)
    print(wb.sheetnames)
    ws = wb['Sheet2']
    print('修改后的工作表', ws)
    # print('之前的值： ', ws['a3'].value)

    # ws['a3'] = '小黑-修改后'
    # print(ws['a3'].value)

    # wb.save('../static/excels/excel_1.xlsx')
main()